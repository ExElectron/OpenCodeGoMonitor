# -*- coding: utf-8 -*-
"""数据导出：CSV / 多 Sheet Excel（后台线程执行，不阻塞 GUI）。

Excel Sheet 结构：
  Sheet1 Raw_Data       原始明细（筛选或全部）
  Sheet2 Summary_Charts 全局汇总指标 + 汇总统计表
  Sheet3 By_API_Key     按 API Key 聚合
  Sheet4 By_Model       按模型聚合
  Sheet5 说明            字段口径 / 单位 / 生成信息
"""
import os
import time

import pandas as pd
from PySide6.QtCore import QThread, Signal

from ocgmon import APP_NAME, APP_VERSION, COST_DIVISOR

# 导出列（与数据库字段对应）
RAW_COLUMNS = [
    ("timestamp_local", "时间(本地)"),
    ("model_name", "模型"),
    ("provider", "提供方"),
    ("key_id", "API Key"),
    ("api_key_masked", "API Key(脱敏)"),
    ("input_raw", "输入Tokens"),
    ("cache_read_tokens", "缓存读"),
    ("cache_write5m_tokens", "缓存写5m"),
    ("cache_write1h_tokens", "缓存写1h"),
    ("prompt_tokens", "总输入Tokens"),
    ("completion_tokens", "输出Tokens"),
    ("reasoning_tokens", "推理Tokens"),
    ("total_tokens", "总Tokens"),
    ("cost", "成本($)"),
    ("tag", "标签"),
    ("id", "记录ID"),
    ("timestamp", "时间(UTC)"),
]

MONEY_FMT = "#,##0.0000"
TOKEN_FMT = "#,##0"


def _df_from_rows(rows: list, columns: list) -> pd.DataFrame:
    data = {}
    for col, label in columns:
        data[label] = [r.get(col) for r in rows]
    return pd.DataFrame(data)


def export_csv(rows: list, path: str) -> str:
    df = _df_from_rows(rows, RAW_COLUMNS)
    df.to_csv(path, index=False, encoding="utf-8-sig")   # BOM 保证 Excel 打开不乱码
    return path


def export_xlsx(rows: list, path: str, meta: dict = None) -> str:
    """导出多 Sheet Excel。rows: 数据库记录列表；meta: {workspace_id, exported_at, filters}"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    money_fmt, token_fmt = MONEY_FMT, TOKEN_FMT

    def _sheet_from_df(ws, df, num_fmt_cols=None):
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for _, r in df.iterrows():
            ws.append([None if pd.isna(v) else v for v in r.tolist()])
        for col in num_fmt_cols or []:
            for cell in ws[col][1:]:
                cell.number_format = money_fmt if col in (14,) else token_fmt
        widths = [max(len(str(c)) if not pd.isna(c) else 0 for c in df[col].tolist()) + 2
                  for col in df.columns]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(w, 8), 40)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        return ws

    # ---- Sheet 1: Raw_Data ----
    ws1 = wb.active
    ws1.title = "Raw_Data"
    df_raw = _df_from_rows(rows, RAW_COLUMNS)
    # 数值列 3-14：Input=4, 缓存=5-7, 总输入=8, 输出=9, 推理=10, 总=11, 成本=12(含0索引)…
    _sheet_from_df(ws1, df_raw, num_fmt_cols=list(range(5, 16)))

    # ---- Sheet 2: Summary_Charts ----
    ws2 = wb.create_sheet("Summary_Charts")
    total_tokens = sum(r.get("total_tokens") or 0 for r in rows)
    total_cost = sum(r.get("cost") or 0 for r in rows)
    total_in = sum(r.get("prompt_tokens") or 0 for r in rows)
    total_out = sum(r.get("completion_tokens") or 0 for r in rows)
    models = {}
    for r in rows:
        m = models.setdefault(r.get("model_name", "unknown"), {"calls": 0, "tokens": 0, "cost": 0.0})
        m["calls"] += 1; m["tokens"] += r.get("total_tokens") or 0; m["cost"] += r.get("cost") or 0
    top_model = max(models, key=lambda k: models[k]["tokens"]) if models else "-"

    summary = [
        ("指标", "数值"),
        ("应用", f"{APP_NAME} v{APP_VERSION}"),
        ("导出时间", meta.get("exported_at", time.strftime("%Y-%m-%d %H:%M:%S")) if meta else time.strftime("%Y-%m-%d %H:%M:%S")),
        ("工作区", meta.get("workspace_id", "-") if meta else "-"),
        ("记录条数", len(rows)),
        ("调用次数", len(rows)),
        ("总输入Tokens(含缓存)", total_in),
        ("总输出Tokens", total_out),
        ("总Tokens", total_tokens),
        ("总成本($)", round(total_cost, 4)),
        ("使用最多模型", top_model),
        ("成本占比最高模型", max(models, key=lambda k: models[k]["cost"]) if models else "-"),
        ("成本单位说明", f"1 美元 = {int(COST_DIVISOR):,} 原始单位；金额保留 4 位小数"),
        ("Token口径说明", "总输入 = inputTokens + cacheRead + cacheWrite5m + cacheWrite1h；总 = 总输入 + output"),
    ]
    for row in summary:
        ws2.append(row)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 36
    for r in range(3, len(summary) + 1):
        ws2.cell(row=r, column=2).number_format = money_fmt if isinstance(ws2.cell(row=r, column=2).value, float) else token_fmt
    # 汇总统计表（按日）
    daily = {}
    for r in rows:
        day = (r.get("timestamp_local") or "")[:10] or "-"
        d = daily.setdefault(day, {"calls": 0, "tokens": 0, "cost": 0.0})
        d["calls"] += 1; d["tokens"] += r.get("total_tokens") or 0; d["cost"] += r.get("cost") or 0
    start_row = len(summary) + 3
    ws2.cell(row=start_row, column=1, value="按日期汇总").font = Font(bold=True)
    daily_rows = [["日期", "调用次数", "总Tokens", "成本($)"]] + \
                 [[d, daily[d]["calls"], daily[d]["tokens"], round(daily[d]["cost"], 4)]
                  for d in sorted(daily)]
    for row in daily_rows:
        ws2.append(row)
    for cell in ws2[start_row + 1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in range(start_row + 2, start_row + 2 + len(daily)):
        ws2.cell(row=r, column=4).number_format = money_fmt
        ws2.cell(row=r, column=3).number_format = token_fmt

    # ---- Sheet 3: By_API_Key ----
    ws3 = wb.create_sheet("By_API_Key")
    keys = {}
    for r in rows:
        k = keys.setdefault(r.get("key_id", "-"), {"calls": 0, "tokens": 0, "cost": 0.0})
        k["calls"] += 1; k["tokens"] += r.get("total_tokens") or 0; k["cost"] += r.get("cost") or 0
    tot_tokens = total_tokens or 1
    tot_cost = total_cost or 0.0001
    df_keys = pd.DataFrame([
        {"API Key": k, "调用次数": v["calls"], "总Tokens": v["tokens"], "总成本($)": round(v["cost"], 4),
         "Token占比": f"{v['tokens'] / tot_tokens * 100:.2f}%", "成本占比": f"{v['cost'] / tot_cost * 100:.2f}%"}
        for k, v in sorted(keys.items(), key=lambda kv: -kv[1]["cost"])
    ])
    _sheet_from_df(ws3, df_keys, num_fmt_cols=[2, 3])

    # ---- Sheet 4: By_Model ----
    ws4 = wb.create_sheet("By_Model")
    df_models = pd.DataFrame([
        {"模型": m, "调用次数": v["calls"], "总Tokens": v["tokens"], "总成本($)": round(v["cost"], 4),
         "Token占比": f"{v['tokens'] / tot_tokens * 100:.2f}%", "成本占比": f"{v['cost'] / tot_cost * 100:.2f}%"}
        for m, v in sorted(models.items(), key=lambda kv: -kv[1]["cost"])
    ])
    _sheet_from_df(ws4, df_models, num_fmt_cols=[2, 3])

    # ---- Sheet 5: 说明 ----
    ws5 = wb.create_sheet("说明")
    notes = [
        ("字段口径与单位", ""),
        ("成本单位", f"接口原始 cost 单位 = 1/1e8 美元；本表已换算为美元，保留 4 位小数。"),
        ("总输入Tokens", "inputTokens + cacheReadTokens + cacheWrite5mTokens + cacheWrite1hTokens"),
        ("总Tokens", "总输入Tokens + outputTokens"),
        ("时间", "时间(本地) = UTC +8 小时（中国标准时间）；时间(UTC) 为接口原始时间"),
        ("防重说明", "记录以 usg_ 唯一 ID 为主键，重复导入自动忽略"),
        ("导出说明", f"由 {APP_NAME} v{APP_VERSION} 生成"),
    ]
    for row in notes:
        ws5.append(row)
    ws5.column_dimensions["A"].width = 24
    ws5.column_dimensions["B"].width = 90

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 后台导出线程
# ---------------------------------------------------------------------------
class ExportWorker(QThread):
    progress = Signal(str, int)          # stage, percent(0-100)
    finished_ok = Signal(str)            # file path
    failed = Signal(str)                 # message

    def __init__(self, rows: list, path: str, fmt: str, meta: dict = None, parent=None):
        super().__init__(parent)
        self._rows, self._path, self._fmt = rows, path, fmt
        self._meta = meta or {}

    def run(self):
        try:
            self.progress.emit("正在准备数据…", 10)
            if self._fmt == "csv":
                export_csv(self._rows, self._path)
            else:
                export_xlsx(self._rows, self._path, self._meta)
            self.progress.emit("导出完成", 100)
            self.finished_ok.emit(self._path)
        except Exception as e:
            self.failed.emit(f"导出失败：{e}")
